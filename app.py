"""
SPARROW Installation Tracker — Streamlit App (v2 Fluent Design)

Run:  streamlit run app.py
"""

import os
import sys
import json
from datetime import datetime, date, timedelta

import streamlit as st
import pandas as pd

sys.path.insert(0, os.path.dirname(__file__))

from db import (
    init_db, get_all_projects, get_project, update_project,
    add_history, get_project_history, get_recent_history,
    add_contact, get_contacts, add_raw_input,
    get_active_nudges, get_status_summary,
    get_stale_projects, get_deadline_approaching,
    get_timeline_rows,
)
from config import (
    VALID_STATUSES, CLOSED_STATUSES, TEAM_MEMBERS,
    AZURE_OPENAI_ENDPOINT, AZURE_OPENAI_DEPLOYMENT,
    STALENESS_THRESHOLDS, IMAP_HOST,
    PHASE_STATUS_COLORS, PHASE_LIGHT_FILL_STATUSES,
)
from theme import (
    inject_theme, render_hero, render_floating_ask,
    status_pill_html, severity_badge_html, confidence_badge_html,
    metric_card_html,
    timeline_entry_html, COLORS,
)

init_db()

# ── Bulk Excel import/export helpers ──────────────────────────────────────────

_BULK_EDITABLE_FIELDS = [
    "continent", "country", "location", "partner_org",
    "status", "team_owner", "target_date", "timeline_label",
    "deployment_type", "hardware", "estimated_cost", "blocker", "notes",
    "track_name",
]
_BULK_EXPORT_COLUMNS = ["project_id", "item_type"] + _BULK_EDITABLE_FIELDS


def _bulk_norm(v):
    """Normalize values for equality comparison between DB and uploaded cells."""
    if v is None:
        return ""
    if isinstance(v, float) and v != v:  # NaN
        return ""
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    # Canonicalize numerics so 25000.0 and 25000 compare equal
    # (openpyxl strips trailing zeros when Excel saves integers).
    if isinstance(v, bool):
        return str(v)
    if isinstance(v, (int, float)):
        return f"{float(v):g}"
    return str(v).strip()


def _export_projects_xlsx(projects: list) -> bytes:
    import io
    import openpyxl
    from openpyxl.utils import get_column_letter
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "projects"
    ws.append(_BULK_EXPORT_COLUMNS)
    for p in projects:
        ws.append([p.get(c) for c in _BULK_EXPORT_COLUMNS])
    for col_idx, col_name in enumerate(_BULK_EXPORT_COLUMNS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(14, len(col_name) + 2)
    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _parse_bulk_upload(uploaded_file):
    """
    Read the xlsx, diff rows against current DB state, return (diff, errors).
      diff: [{"project_id", "field", "old", "new"}]
      errors: [str]  — rows/fields that were skipped, with reason.
    """
    import openpyxl
    wb = openpyxl.load_workbook(uploaded_file, data_only=True)
    ws = wb["projects"] if "projects" in wb.sheetnames else wb.active
    header = [c.value for c in ws[1]]
    if "project_id" not in header:
        raise ValueError("Sheet is missing a 'project_id' column")

    diff = []
    errors = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row_dict = dict(zip(header, row))
        pid = row_dict.get("project_id")
        if not pid:
            continue
        current = get_project(pid)
        if not current:
            errors.append(f"Row {row_idx}: project_id '{pid}' not found — skipping row")
            continue

        for field in _BULK_EDITABLE_FIELDS:
            if field not in row_dict:
                continue
            raw_new = row_dict[field]
            old = current.get(field)

            if _bulk_norm(old) == _bulk_norm(raw_new):
                continue

            # Coerce / validate
            new = raw_new
            if field == "status":
                if new and str(new) not in VALID_STATUSES:
                    errors.append(f"Row {row_idx} ({pid}): invalid status '{new}' — skipping field")
                    continue
            elif field == "target_date":
                if isinstance(new, datetime):
                    new = new.date().isoformat()
                elif isinstance(new, date):
                    new = new.isoformat()
                elif new not in (None, ""):
                    try:
                        date.fromisoformat(str(new))
                        new = str(new)
                    except ValueError:
                        errors.append(f"Row {row_idx} ({pid}): target_date '{raw_new}' must be YYYY-MM-DD — skipping field")
                        continue
            elif field == "estimated_cost":
                if new not in (None, ""):
                    try:
                        new = float(new)
                    except (ValueError, TypeError):
                        errors.append(f"Row {row_idx} ({pid}): estimated_cost '{raw_new}' must be numeric — skipping field")
                        continue

            if new == "":
                new = None

            diff.append({"project_id": pid, "field": field, "old": old, "new": new})

    return diff, errors


# ── Page Config ───────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="SPARROW Tracker",
    page_icon="\U0001F426",
    layout="wide",
    initial_sidebar_state="expanded",
)

inject_theme()

llm_available = bool(AZURE_OPENAI_ENDPOINT)

# ── Sidebar ───────────────────────────────────────────────────────────────────

with st.sidebar:
    st.markdown(
        '<div style="display:flex;align-items:center;gap:10px;padding:4px 0 16px">'
        '<span style="font-size:28px">🐦</span>'
        '<div><div style="font-size:18px;font-weight:700;color:#242424">SPARROW</div>'
        '<div style="font-size:11px;color:#616161">Installation Tracker</div></div></div>',
        unsafe_allow_html=True,
    )

    # Apply any pending navigation before the radio is instantiated.
    # (Can't write to nav_page after the widget exists on the same run.)
    if "_pending_nav" in st.session_state:
        st.session_state["nav_page"] = st.session_state.pop("_pending_nav")

    page = st.radio(
        "Navigate",
        ["Dashboard", "Submit Update", "Timeline", "Project Details", "Sprints", "Reports", "Settings"],
        index=0,
        label_visibility="collapsed",
        key="nav_page",
    )

    st.markdown("---")

    # Ask SPARROW section in sidebar
    st.markdown(
        '<div style="font-size:13px;font-weight:600;color:#242424;margin-bottom:8px">'
        '💬 Ask SPARROW</div>',
        unsafe_allow_html=True,
    )
    quick_queries = [
        "What's blocked or at risk?",
        "Recent changes",
        "FY26 deadlines",
        "Robin dependencies",
    ]
    for q in quick_queries:
        if st.button(q, key=f"sq_{q}", use_container_width=True):
            st.session_state["ask_question"] = q
            st.session_state["ask_pending"] = True

    ask_input = st.text_input(
        "Ask anything...",
        key="ask_sidebar_input",
        label_visibility="collapsed",
        placeholder="Ask about your projects...",
    )
    if ask_input and ask_input != st.session_state.get("_last_ask", ""):
        st.session_state["ask_question"] = ask_input
        st.session_state["ask_pending"] = True
        st.session_state["_last_ask"] = ask_input

    # Process the input: may be a question OR an update (e.g. "assign Carl to Salonga").
    if st.session_state.get("ask_pending") and llm_available:
        question = st.session_state.get("ask_question", "")
        if question:
            with st.spinner("Thinking..."):
                from llm import parse_input, answer_question
                parsed = parse_input(question, "Ask SPARROW")

            intent = parsed.get("input_type", "question")
            st.session_state["ask_pending"] = False

            if intent in ("update", "new_project") and parsed.get("proposed_changes"):
                # Route through the Submit Update page so the user can review/approve
                # using the existing confirmation UI.
                st.session_state["pending_result"] = parsed
                st.session_state["pending_text"] = question
                st.session_state["pending_type"] = "ask_sparrow"
                st.session_state["pending_by"] = "Ask SPARROW"
                st.session_state["ask_answer"] = (
                    "Proposed changes ready for review on the Submit Update page."
                )
                st.session_state["_pending_nav"] = "Submit Update"
                add_raw_input(question, input_type="update")
                st.rerun()
            else:
                # Either a question or an update that couldn't produce structured changes.
                answer = parsed.get("question_answer")
                if not answer:
                    answer = answer_question(question)
                st.session_state["ask_answer"] = answer
                add_raw_input(question, input_type="question")

    if st.session_state.get("ask_answer"):
        st.markdown(
            f'<div style="background:#eff6fc;border:1px solid #c7e0f4;border-radius:8px;'
            f'padding:12px;font-size:13px;max-height:300px;overflow-y:auto">'
            f'{st.session_state["ask_answer"]}</div>',
            unsafe_allow_html=True,
        )
        if st.button("Clear", key="clear_ask"):
            st.session_state["ask_answer"] = None
            st.session_state["ask_question"] = ""
            st.rerun()

    if not llm_available:
        st.caption("⚠️ LLM not configured")

    st.markdown("---")
    all_projects_count = len(get_all_projects())
    st.caption(f"{all_projects_count} Projects · {date.today().strftime('%B %d, %Y')}")


# ══════════════════════════════════════════════════════════════════════════════
# DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════

if page == "Dashboard":

    # ── Header ────────────────────────────────────────────────────────────
    st.markdown("## Dashboard")

    # ── Hero Banner ───────────────────────────────────────────────────────
    projects = get_all_projects()
    continents = len(set(p["continent"] for p in projects))
    countries = len(set(p["country"] for p in projects))
    active = sum(1 for p in projects if (p["status"] or "").startswith("Active"))
    completed = sum(1 for p in projects if p["status"] == "Complete")
    render_hero(len(projects), continents, countries, active, completed)

    # ── Needs Attention (collapsible, click an item to open it) ───────────
    stale = get_stale_projects()
    deadline = get_deadline_approaching()
    attention_items = stale + [d for d in deadline if d["project_id"] not in {s["project_id"] for s in stale}]
    overdue = [d for d in deadline if d.get("days_until_deadline", 1) <= 0]

    if attention_items:
        expander_label = f"⚠️ Needs Attention ({len(attention_items)})"
        if overdue:
            expander_label = f"🚨 Needs Attention — {len(overdue)} overdue, {len(attention_items)} total"
        with st.expander(expander_label, expanded=bool(overdue)):
            for item in attention_items:
                if "days_since_update" in item:
                    detail = f"{item['days_since_update']}d since last update"
                    sev_color = "#d13438" if item["days_since_update"] > 30 else "#ca5010"
                else:
                    d = item.get("days_until_deadline", 0)
                    detail = f"OVERDUE by {-d}d" if d < 0 else f"{d}d to deadline"
                    sev_color = "#d13438" if d <= 0 else ("#ca5010" if d <= 14 else "#8a8886")
                blocker = item.get("blocker") or ""
                row = st.columns([6, 2, 2, 2])
                with row[0]:
                    if st.button(
                        f"{item['location']} · {item.get('country', '')}",
                        key=f"attn_{item['project_id']}",
                        use_container_width=True,
                    ):
                        st.session_state["_pending_nav"] = "Project Details"
                        st.session_state["_pending_project_pid"] = item["project_id"]
                        st.rerun()
                with row[1]:
                    st.markdown(
                        f'<div style="padding:8px 0;font-size:12px;color:#616161">{item["status"]}</div>',
                        unsafe_allow_html=True,
                    )
                with row[2]:
                    st.markdown(
                        f'<div style="padding:8px 0;font-size:12px;color:{sev_color};font-weight:600">{detail}</div>',
                        unsafe_allow_html=True,
                    )
                with row[3]:
                    if blocker:
                        st.markdown(
                            f'<div style="padding:8px 0;font-size:11px;color:#8a8886" title="{blocker}">{blocker[:35]}{"…" if len(blocker) > 35 else ""}</div>',
                            unsafe_allow_html=True,
                        )

    # ── Stat Cards (click to filter the table below) ─────────────────────
    summary = get_status_summary()
    sorted_statuses = sorted(summary.items(), key=lambda x: -x[1])

    bar_color_map = {
        "Scoping": COLORS["neutral"],
        "Active - Waiting on Partner": COLORS["warning"],
        "Active - Waiting on Us": COLORS["primary"],
        "Complete": COLORS["success"],
        "Descoped": COLORS["neutral"],
    }

    if "dashboard_status_filter" not in st.session_state:
        st.session_state["dashboard_status_filter"] = []

    # Scoped CSS so only these buttons get the card treatment.
    st.markdown(
        """
        <style>
        div[data-dashboard-stats] div.stButton > button {
            background: #fff; border: 1px solid #edebe9; border-radius: 8px;
            box-shadow: 0 1px 2px rgba(0,0,0,0.04);
            padding: 18px 18px 14px; min-height: 92px;
            text-align: left; width: 100%;
            font-weight: 500; color: #242424;
            transition: all 0.15s ease;
            line-height: 1.25;
        }
        div[data-dashboard-stats] div.stButton > button:hover {
            transform: translateY(-2px);
            box-shadow: 0 2px 8px rgba(0,0,0,0.08);
            border-color: #d2d0ce;
        }
        div[data-dashboard-stats] div.stButton > button:focus {
            outline: none; box-shadow: 0 0 0 2px rgba(0,120,212,0.25);
        }
        div[data-dashboard-stats] div.stButton > button[kind="primary"] {
            background: #eff6fc; border-color: #0078d4; color: #242424;
        }
        </style>
        <div data-dashboard-stats></div>
        """,
        unsafe_allow_html=True,
    )

    stat_cols = st.columns(len(sorted_statuses))
    for i, (status, count) in enumerate(sorted_statuses):
        with stat_cols[i]:
            is_active = st.session_state["dashboard_status_filter"] == [status]
            if st.button(
                f"{count}\n\n{status}",
                key=f"stat_{status}",
                use_container_width=True,
                type="primary" if is_active else "secondary",
                help=f"Click to filter table to {status} projects (click again to clear).",
            ):
                st.session_state["dashboard_status_filter"] = (
                    [] if is_active else [status]
                )
                st.rerun()

    st.markdown(
        f'<div style="height:3px;margin:-6px 0 22px;background:linear-gradient(to right,'
        + ','.join(f'{bar_color_map.get(s, COLORS["neutral"])} {i*100/len(sorted_statuses):.1f}%,'
                   f'{bar_color_map.get(s, COLORS["neutral"])} {(i+1)*100/len(sorted_statuses):.1f}%'
                   for i, (s, _) in enumerate(sorted_statuses))
        + ')"></div>',
        unsafe_allow_html=True,
    )

    # ── All Projects Table ────────────────────────────────────────────────
    st.markdown('<div class="section-title">All Projects</div>',
                unsafe_allow_html=True)

    filter_cols = st.columns(4)
    with filter_cols[0]:
        status_filter = st.multiselect(
            "Status", VALID_STATUSES,
            key="dashboard_status_filter",
        )
    with filter_cols[1]:
        continent_list = sorted(set(p["continent"] for p in projects))
        continent_filter = st.multiselect("Continent", continent_list, default=[])
    with filter_cols[2]:
        owners = sorted(set(p.get("team_owner") or "Unassigned" for p in projects))
        owner_filter = st.multiselect("Owner", owners, default=[])
    with filter_cols[3]:
        search = st.text_input("Search", "", label_visibility="visible")

    filtered = projects
    if status_filter:
        filtered = [p for p in filtered if p["status"] in status_filter]
    if continent_filter:
        filtered = [p for p in filtered if p["continent"] in continent_filter]
    if owner_filter:
        filtered = [p for p in filtered if (p.get("team_owner") or "Unassigned") in owner_filter]
    if search:
        sl = search.lower()
        filtered = [p for p in filtered if sl in json.dumps(p).lower()]

    df = pd.DataFrame(filtered)
    # Columns the user can edit directly in the grid. project_id and
    # last_updated stay read-only so identity and audit fields aren't touched.
    editable_cols = ["status", "team_owner", "target_date", "estimated_cost", "partner_org"]
    display_cols = [
        "project_id", "location", "partner_org",
        "status", "team_owner", "target_date",
        "estimated_cost", "last_updated",
    ]
    display_cols = [c for c in display_cols if c in df.columns]

    if df.empty:
        st.info("No projects match the current filters.")
    else:
        edit_col, save_col = st.columns([6, 1])
        with save_col:
            save_clicked = st.button("Save changes", type="primary", use_container_width=True)

        owner_options = sorted(
            {p.get("team_owner") for p in projects if p.get("team_owner")}
            | set(TEAM_MEMBERS)
        )

        edited_df = st.data_editor(
            df[display_cols],
            use_container_width=True,
            hide_index=True,
            num_rows="fixed",
            disabled=[c for c in display_cols if c not in editable_cols],
            key=f"dashboard_editor_{len(filtered)}",
            column_config={
                "project_id": st.column_config.TextColumn("ID", width="small"),
                "location": st.column_config.TextColumn("Location"),
                "partner_org": st.column_config.TextColumn("Partner"),
                "status": st.column_config.SelectboxColumn(
                    "Status", options=VALID_STATUSES, required=True
                ),
                "team_owner": st.column_config.SelectboxColumn(
                    "Owner", options=owner_options
                ),
                "estimated_cost": st.column_config.NumberColumn("Cost (USD)", format="$%.0f"),
                "target_date": st.column_config.TextColumn("Target Date", help="YYYY-MM-DD"),
                "last_updated": st.column_config.TextColumn("Last Updated", disabled=True),
            },
        )

        if save_clicked:
            saved = 0
            errors = []
            orig_by_pid = {p["project_id"]: p for p in filtered}
            for _, row in edited_df.iterrows():
                pid = row["project_id"]
                orig = orig_by_pid.get(pid)
                if not orig:
                    continue
                updates = {}
                for col in editable_cols:
                    if col not in edited_df.columns:
                        continue
                    new_val = row[col]
                    old_val = orig.get(col)
                    # pandas hands us NaN for empty cells — normalize.
                    if pd.isna(new_val):
                        new_val = None
                    if str(old_val or "") != str(new_val or ""):
                        updates[col] = new_val
                if not updates:
                    continue
                try:
                    changes = update_project(pid, updates, "Dashboard edit")
                    if changes:
                        fields = ", ".join(changes.keys())
                        add_history(
                            pid, changes,
                            source_text="Dashboard grid edit",
                            source_type="manual",
                            updated_by="Dashboard edit",
                            llm_summary=f"Edited {fields} via Dashboard",
                        )
                        saved += 1
                except Exception as e:
                    errors.append(f"{pid}: {e}")

            if saved:
                st.success(f"Saved {saved} project(s).")
            if errors:
                st.error("Some rows failed:\n" + "\n".join(errors))
            if saved and not errors:
                st.rerun()

    # ── Active Nudges ─────────────────────────────────────────────────────
    nudges = get_active_nudges()
    if nudges:
        with st.expander(f"Active Nudges ({len(nudges)})"):
            for n in nudges:
                st.markdown(
                    f"{severity_badge_html(n['severity'])} **{n['project_id']}** — {n['message'][:150]}",
                    unsafe_allow_html=True,
                )

    # ── Bulk edit via Excel ──────────────────────────────────────────────
    with st.expander("📥 Bulk edit via Excel"):
        st.caption(
            "Export the project list to Excel, edit offline, re-upload to apply changes. "
            f"Editable columns: {', '.join(_BULK_EDITABLE_FIELDS)}. "
            "project_id is the key — rows with unknown project_ids are skipped."
        )

        be_cols = st.columns([2, 5])
        with be_cols[0]:
            xlsx_bytes = _export_projects_xlsx(get_all_projects())
            st.download_button(
                "📤  Export to Excel",
                data=xlsx_bytes,
                file_name=f"sparrow_projects_{date.today().isoformat()}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="bulk_export_btn",
            )

        st.markdown("<hr style='margin:10px 0;border:none;border-top:1px solid #edebe9'>",
                    unsafe_allow_html=True)

        uploaded = st.file_uploader(
            "Upload edited Excel (.xlsx)",
            type=["xlsx"], key="bulk_upload_file",
            help="File should have a 'projects' sheet (or a single sheet). "
                 "Must include a 'project_id' column.",
        )

        if uploaded is not None:
            try:
                diff, errors = _parse_bulk_upload(uploaded)
            except Exception as e:
                st.error(f"Could not read file: {e}")
                diff, errors = [], []

            if errors:
                with st.expander(f"⚠️ {len(errors)} warning(s)"):
                    for msg in errors:
                        st.caption(f"• {msg}")

            if not diff:
                st.info("No changes detected against current DB state.")
            else:
                st.markdown(f"**{len(diff)} change(s) detected.** Review before applying.")
                preview_df = pd.DataFrame(diff)[["project_id", "field", "old", "new"]]
                st.dataframe(preview_df, use_container_width=True, hide_index=True, height=280)

                apply_cols = st.columns([5, 2])
                with apply_cols[1]:
                    apply_clicked = st.button(
                        f"Apply {len(diff)} change(s)",
                        type="primary", use_container_width=True,
                        key="bulk_apply_btn",
                    )
                if apply_clicked:
                    applied = 0
                    failures = []
                    # Group changes by project_id so each project gets one history row.
                    by_pid = {}
                    for c in diff:
                        by_pid.setdefault(c["project_id"], {})[c["field"]] = c["new"]

                    for pid, updates in by_pid.items():
                        try:
                            changes = update_project(pid, updates, "Bulk Excel upload")
                            if changes:
                                add_history(
                                    pid, changes,
                                    source_text=f"Bulk xlsx upload: {uploaded.name}",
                                    source_type="manual",
                                    updated_by="Bulk Excel upload",
                                    llm_summary=f"Bulk edit: {', '.join(changes.keys())}",
                                )
                                applied += 1
                        except Exception as e:
                            failures.append(f"{pid}: {e}")

                    if applied:
                        st.success(f"Applied changes to {applied} project(s).")
                    if failures:
                        st.error("Some projects failed:\n" + "\n".join(failures))
                    if applied and not failures:
                        st.rerun()


# ══════════════════════════════════════════════════════════════════════════════
# SUBMIT UPDATE
# ══════════════════════════════════════════════════════════════════════════════

elif page == "Submit Update":
    st.markdown("## Submit Update")
    st.markdown(
        "Paste an email, Teams message, meeting notes, or plain English. "
        "The AI will identify projects and propose changes."
    )

    input_col, context_col = st.columns([3, 2])

    with input_col:
        with st.form("submit_update_form", clear_on_submit=False):
            text = st.text_area(
                "Paste your update here", height=280,
                placeholder="e.g., 'Paulo confirmed Salonga shipment arrived in Kinshasa yesterday. "
                "Still waiting on customs clearance.'\n\nPaste emails, Teams messages, meeting notes, "
                "or plain English — the AI will figure out the rest.",
                key="submit_update_text",
            )
            submitted = st.form_submit_button(
                "Send to SPARROW", type="primary",
                disabled=not llm_available, use_container_width=True,
            )

        if submitted:
            if not text.strip():
                st.warning("Paste some text before submitting.")
            else:
                with st.spinner("Analyzing..."):
                    from llm import parse_input
                    result = parse_input(text, "auto")
                st.session_state["pending_result"] = result
                st.session_state["pending_text"] = text
                st.session_state["pending_type"] = "update"
                st.session_state["pending_by"] = "System"

        # ── Results ───────────────────────────────────────────────────────
        if "pending_result" in st.session_state:
            result = st.session_state["pending_result"]

            if result.get("input_type") == "question":
                st.info("This looks like a question:")
                st.markdown(result.get("question_answer") or result.get("llm_summary", ""))
                if st.button("Clear"):
                    del st.session_state["pending_result"]
                    st.rerun()

            elif result.get("input_type") == "unclear":
                st.warning("Could not clearly identify which project this relates to.")
                st.json(result)
                if st.button("Clear"):
                    del st.session_state["pending_result"]
                    st.rerun()

            else:
                # Summary banner
                matches = result.get("matched_projects", [])
                all_high = all(m.get("match_confidence") == "high" for m in matches)
                banner_color = "#dff6dd" if all_high else "#fff4ce"
                banner_icon = "✓" if all_high else "?"
                st.markdown(
                    f'<div style="background:{banner_color};border-radius:8px;padding:12px 16px;'
                    f'margin:12px 0;font-size:14px">'
                    f'{banner_icon} Matched {len(matches)} project(s) '
                    f'{"with high confidence" if all_high else "— review confidence below"}</div>',
                    unsafe_allow_html=True,
                )

                st.markdown(f"*{result.get('llm_summary', '')}*")

                # Matched projects
                for match in matches:
                    conf = match.get("match_confidence", "?")
                    st.markdown(
                        f"**{match['project_id']}** — {confidence_badge_html(conf)} — "
                        f"{match.get('match_reason', '')}",
                        unsafe_allow_html=True,
                    )

                def _clean_rows(df):
                    rows = df.to_dict("records")
                    for r in rows:
                        for k, v in list(r.items()):
                            if isinstance(v, float) and pd.isna(v):
                                r[k] = None
                            elif hasattr(v, "item") and not isinstance(v, str):
                                try:
                                    r[k] = v.item()
                                except (ValueError, AttributeError):
                                    pass
                    return rows

                # Proposed project-level changes (editable)
                changes = result.get("proposed_changes", [])
                edited_changes = changes
                if changes:
                    st.markdown("**Proposed project changes** — edit values or delete rows before approving:")
                    change_df = pd.DataFrame(changes)
                    for col in ("project_id", "field", "new_value", "evidence"):
                        if col not in change_df.columns:
                            change_df[col] = None
                    change_df = change_df[["project_id", "field", "new_value", "evidence"]]
                    edited_changes_df = st.data_editor(
                        change_df,
                        num_rows="dynamic",
                        use_container_width=True, hide_index=True,
                        key="edit_proposed_changes",
                        column_config={
                            "project_id": st.column_config.TextColumn("Project", disabled=True),
                            "field": st.column_config.TextColumn("Field", disabled=True),
                            "new_value": st.column_config.TextColumn("New value"),
                            "evidence": st.column_config.TextColumn("Evidence", disabled=True),
                        },
                    )
                    edited_changes = _clean_rows(edited_changes_df)

                # Proposed phase-level changes (editable)
                phase_changes = result.get("phase_changes", [])
                edited_phase_changes = []
                if phase_changes:
                    st.markdown("**Proposed phase changes** — edit dates/status or delete rows before approving:")
                    rows = []
                    for pc in phase_changes:
                        fu = pc.get("field_updates") or {}
                        rows.append({
                            "project_id": pc.get("project_id"),
                            "phase_id": pc.get("phase_id"),
                            "action": pc.get("action"),
                            "name": fu.get("name"),
                            "start_date": fu.get("start_date"),
                            "end_date": fu.get("end_date"),
                            "status": fu.get("status"),
                            "evidence": pc.get("evidence"),
                        })
                    from config import PHASE_STATUSES
                    phase_action_options = ["update", "create", "delete"]
                    edited_phase_df = st.data_editor(
                        pd.DataFrame(rows),
                        num_rows="dynamic",
                        use_container_width=True, hide_index=True,
                        key="edit_phase_changes",
                        column_config={
                            "project_id": st.column_config.TextColumn("Project", disabled=True),
                            "phase_id": st.column_config.NumberColumn("Phase id", help="Blank for create"),
                            "action": st.column_config.SelectboxColumn(
                                "Action", options=phase_action_options, default="update",
                            ),
                            "name": st.column_config.TextColumn("Name"),
                            "start_date": st.column_config.TextColumn("Start (YYYY-MM-DD)"),
                            "end_date": st.column_config.TextColumn("End (YYYY-MM-DD)"),
                            "status": st.column_config.SelectboxColumn(
                                "Status", options=[None] + list(PHASE_STATUSES),
                            ),
                            "evidence": st.column_config.TextColumn("Evidence", disabled=True),
                        },
                    )
                    for r in _clean_rows(edited_phase_df):
                        edited_phase_changes.append({
                            "project_id": r.get("project_id"),
                            "phase_id": r.get("phase_id"),
                            "action": r.get("action") or "update",
                            "field_updates": {
                                "name": r.get("name"),
                                "start_date": r.get("start_date"),
                                "end_date": r.get("end_date"),
                                "status": r.get("status"),
                            },
                            "evidence": r.get("evidence"),
                        })

                # New contacts (editable)
                contacts = result.get("new_contacts", [])
                edited_contacts = contacts
                if contacts:
                    st.markdown("**New contacts detected** — edit or delete rows before approving:")
                    contact_df = pd.DataFrame(contacts)
                    for col in ("name", "organization", "role", "email", "phone", "linked_project"):
                        if col not in contact_df.columns:
                            contact_df[col] = None
                    contact_df = contact_df[["name", "organization", "role", "email", "phone", "linked_project"]]
                    edited_contact_df = st.data_editor(
                        contact_df,
                        num_rows="dynamic",
                        use_container_width=True, hide_index=True,
                        key="edit_new_contacts",
                    )
                    edited_contacts = _clean_rows(edited_contact_df)

                # Action buttons
                act_cols = st.columns([2, 2, 1])
                with act_cols[0]:
                    if st.button("Approve All", type="primary", use_container_width=True):
                        from db import apply_phase_change
                        history_ids = []
                        applied_changes = [c for c in edited_changes
                                           if c.get("project_id") and c.get("field")]
                        for change in applied_changes:
                            pid = change["project_id"]
                            field_changes = update_project(
                                pid, {change["field"]: change.get("new_value")},
                                st.session_state["pending_by"],
                            )
                            if field_changes:
                                hid = add_history(
                                    pid, field_changes,
                                    source_text=st.session_state["pending_text"],
                                    source_type=st.session_state["pending_type"],
                                    updated_by=st.session_state["pending_by"],
                                    llm_summary=result.get("llm_summary"),
                                )
                                history_ids.append(hid)

                        # Apply phase-level changes, one history row per phase change
                        applied_phase_changes = [pc for pc in edited_phase_changes
                                                 if pc.get("project_id")]
                        for pc in applied_phase_changes:
                            pid = pc["project_id"]
                            summary = apply_phase_change(pid, pc)
                            if summary.get("action") == "noop":
                                continue
                            hid = add_history(
                                pid,
                                {"phase": {"old": f"phase_id={summary.get('phase_id')}",
                                           "new": json.dumps(summary)}},
                                source_text=st.session_state["pending_text"],
                                source_type=st.session_state["pending_type"],
                                updated_by=st.session_state["pending_by"],
                                llm_summary=f"Phase {summary.get('action')}: {pc.get('evidence') or ''}"[:250],
                            )
                            history_ids.append(hid)

                        applied_contacts = [c for c in edited_contacts if c.get("name")]
                        for c in applied_contacts:
                            add_contact(
                                name=c.get("name", ""),
                                organization=c.get("organization"),
                                role=c.get("role"),
                                email=c.get("email"),
                                phone=c.get("phone"),
                                linked_projects=[c.get("linked_project")] if c.get("linked_project") else [],
                            )

                        add_raw_input(
                            st.session_state["pending_text"],
                            submitted_by=st.session_state["pending_by"],
                            input_type="update",
                            history_ids=history_ids,
                        )
                        total_applied = len(applied_changes) + len(applied_phase_changes)
                        st.success(f"Applied {total_applied} change(s) "
                                   f"({len(applied_changes)} project, {len(applied_phase_changes)} phase).")
                        del st.session_state["pending_result"]
                        st.rerun()

                with act_cols[2]:
                    if st.button("Reject", use_container_width=True):
                        add_raw_input(
                            st.session_state["pending_text"],
                            submitted_by=st.session_state["pending_by"],
                            input_type="update",
                            history_ids=[],
                        )
                        del st.session_state["pending_result"]
                        st.rerun()

    with context_col:
        # Recently Updated feed
        st.markdown(
            '<div style="background:#fff;border:1px solid #edebe9;border-radius:8px;box-shadow:0 1px 2px rgba(0,0,0,0.04);overflow:hidden">'
            '<div style="padding:14px 16px;border-bottom:1px solid #edebe9;font-size:14px;font-weight:600">'
            'Recently Updated</div><div style="padding:8px 16px">',
            unsafe_allow_html=True,
        )
        recent = get_recent_history(days=14, limit=5)
        if recent:
            items_html = ""
            for h in recent:
                pid = h.get("project_id", "")
                summary_text = h.get("llm_summary") or "Update"[:60]
                ts = h.get("timestamp", "")[:10]
                items_html += (
                    f'<div style="padding:8px 0;border-bottom:1px solid #f3f2f1;font-size:13px">'
                    f'<span style="display:inline-block;padding:1px 6px;background:#eff6fc;'
                    f'border-radius:4px;font-size:11px;color:#0078d4;font-weight:600;margin-right:6px">'
                    f'{pid}</span>{summary_text}'
                    f'<div style="font-size:11px;color:#8a8886;margin-top:2px">{ts}</div></div>'
                )
            st.markdown(items_html + '</div></div>', unsafe_allow_html=True)
        else:
            st.markdown(
                '<div style="padding:16px;color:#8a8886;font-size:13px">No recent updates</div>'
                '</div></div>',
                unsafe_allow_html=True,
            )

        # Quick tip
        st.markdown(
            '<div style="background:#fff;border:1px solid #edebe9;border-radius:8px;box-shadow:0 1px 2px rgba(0,0,0,0.04);'
            'padding:16px;margin-top:12px">'
            '<div style="font-size:14px;font-weight:600;margin-bottom:8px">💡 Quick Tip</div>'
            '<div style="font-size:13px;color:#616161">You can forward emails to your configured '
            'SPARROW inbox and they\'ll be processed automatically. '
            'Configure this in Settings → Email Ingestion.</div></div>',
            unsafe_allow_html=True,
        )

    # Manual update fallback
    with st.expander("Manual update (without AI)"):
        projects = get_all_projects()
        project_options = {f"{p['project_id']} — {p['location']} ({p['partner_org']})": p["project_id"] for p in projects}
        selected = st.selectbox("Select project", list(project_options.keys()), key="manual_proj")
        if selected:
            pid = project_options[selected]
            current = get_project(pid)

            manual_cols = st.columns(2)
            with manual_cols[0]:
                new_status = st.selectbox("Status", VALID_STATUSES, index=VALID_STATUSES.index(current["status"]))
                new_blocker = st.text_input("Blocker / Risk", current.get("blocker") or "")
                new_owner = st.selectbox(
                    "Team Owner", [""] + TEAM_MEMBERS,
                    index=(TEAM_MEMBERS.index(current["team_owner"]) + 1) if current.get("team_owner") in TEAM_MEMBERS else 0,
                )
            with manual_cols[1]:
                new_timeline = st.text_input("Timeline label", current.get("timeline_label") or "")
                new_target = st.text_input("Target date (YYYY-MM-DD)", current.get("target_date") or "")
                new_notes = st.text_area("Notes", current.get("notes") or "", height=100)

            manual_note = st.text_area("Change reason / source", height=80, placeholder="Why is this changing?")
            manual_by = st.selectbox("Updated by", TEAM_MEMBERS + ["Other"], key="manual_by")

            if st.button("Save Manual Update"):
                updates = {}
                if new_status != current["status"]:
                    updates["status"] = new_status
                if new_blocker != (current.get("blocker") or ""):
                    updates["blocker"] = new_blocker or None
                if new_owner != (current.get("team_owner") or ""):
                    updates["team_owner"] = new_owner or None
                if new_timeline != (current.get("timeline_label") or ""):
                    updates["timeline_label"] = new_timeline or None
                if new_target != (current.get("target_date") or ""):
                    updates["target_date"] = new_target or None
                if new_notes != (current.get("notes") or ""):
                    updates["notes"] = new_notes or None

                if updates:
                    field_changes = update_project(pid, updates, manual_by)
                    if field_changes:
                        add_history(
                            pid, field_changes,
                            source_text=manual_note or "Manual update via UI",
                            source_type="manual_note",
                            updated_by=manual_by,
                            llm_summary=f"Manual update: {', '.join(field_changes.keys())} changed",
                        )
                    st.success(f"Updated {len(updates)} field(s) for {pid}.")
                    st.rerun()
                else:
                    st.info("No changes detected.")


# ══════════════════════════════════════════════════════════════════════════════
# TIMELINE
# ══════════════════════════════════════════════════════════════════════════════

elif page == "Timeline":
    st.markdown("## Timeline")
    st.caption(
        "Development tracks and deployments on one Gantt (vis-timeline). "
        "DEVELOPMENT and DEPLOYMENTS sections are collapsible — click the header to fold. "
        "Amber arrows mark active blockers; dashed gray arrows mark resolved dependencies."
    )

    rows = get_timeline_rows(include_closed=True)
    if not rows:
        st.info("No phase data yet. Run `python seed_dev_tracks.py` to seed dev tracks, "
                "and re-open a project so its single synthesized phase shows up.")
    else:
        from timeline_component import render_timeline, build_payload

        today = date.today()

        # ── Filters ──────────────────────────────────────────────────────────
        f_cols = st.columns([2, 2, 2, 3])
        with f_cols[0]:
            types = sorted({r["item_type"] for r in rows})
            sel_types = st.multiselect("Item type", types, default=types)
        with f_cols[1]:
            statuses = sorted({r["phase_status"] for r in rows})
            sel_statuses = st.multiselect("Phase status", statuses, default=statuses)
        with f_cols[2]:
            hide_done = st.checkbox("Hide Done / Cancelled", value=False)
        with f_cols[3]:
            show_deployments = st.checkbox("Show deployment lanes", value=True,
                                           help="Dev tracks are always shown; toggle to hide the ~40 deployment lanes.")

        def lane_for(r):
            if r["item_type"] == "dev_track":
                return r["track_name"] or r["project_id"]
            loc = r["location"] or r["partner_org"] or r["project_id"]
            return f"{r['country']} — {loc}" if r["country"] else loc

        def _normalize_span(r):
            """Return (start, end) strings with both populated, or (None, None) if impossible."""
            start = r["start_date"]
            end = r["end_date"]
            if not start and not end:
                return None, None
            if start and not end:
                end = start
            if end and not start:
                try:
                    end_dt = datetime.fromisoformat(end).date()
                    start = (end_dt - timedelta(days=30)).isoformat()
                except ValueError:
                    return None, None
            return start, end

        # Apply filters to the raw rows once — everything downstream uses filtered_rows.
        filtered_rows = []
        for r in rows:
            if r["item_type"] not in sel_types:
                continue
            if r["phase_status"] not in sel_statuses:
                continue
            if hide_done and r["phase_status"] in ("Done", "Cancelled"):
                continue
            if not show_deployments and r["item_type"] == "deployment":
                continue
            s, _ = _normalize_span(r)
            if not s:
                continue
            filtered_rows.append(r)

        if not filtered_rows:
            st.info("No phases match the current filters.")
        else:
            # ── KPI summary header ──────────────────────────────────────────
            lanes_in_view = list({r["project_id"] for r in filtered_rows})

            def _lane_at_risk(pid):
                for r in filtered_rows:
                    if r["project_id"] != pid:
                        continue
                    if r["phase_status"] in ("At Risk", "Blocked"):
                        return True
                    if r["phase_status"] not in ("Done", "Cancelled") and r["end_date"]:
                        try:
                            if date.fromisoformat(r["end_date"]) < today:
                                return True
                        except ValueError:
                            pass
                return False

            at_risk_lanes = [pid for pid in lanes_in_view if _lane_at_risk(pid)]
            on_track_lanes = [pid for pid in lanes_in_view if pid not in at_risk_lanes]

            LAUNCH_KEYS = {"Launch", "Rollout", "Installed"}
            milestones = []
            for r in filtered_rows:
                if r["phase_status"] in ("Done", "Cancelled"):
                    continue
                if not r["end_date"]:
                    continue
                try:
                    end_dt = date.fromisoformat(r["end_date"])
                except ValueError:
                    continue
                if end_dt < today:
                    continue
                if r["phase_key"] in LAUNCH_KEYS:
                    milestones.append((end_dt, r["phase_name"]))
            milestones.sort()
            next_ms = milestones[0] if milestones else None

            next_ms_text = (
                f'{next_ms[0].strftime("%b %d")} <span style="color:#94a3b8">·</span> '
                f'<span style="color:#64748b">{next_ms[1]}</span>'
                if next_ms else '—'
            )
            kpi_html = (
                '<div style="display:flex;gap:26px;align-items:baseline;flex-wrap:wrap;'
                'padding:14px 20px;margin:0 0 16px;background:#f8fafc;border:1px solid #e2e8f0;'
                'border-radius:8px;font-family:Inter,Segoe UI,Arial,sans-serif;color:#334155">'
                f'<span><span style="color:#64748b;font-size:12px;font-weight:500">Tracks</span> '
                f'<b style="color:#1d4ed8;font-size:16px;margin-left:4px">{len(lanes_in_view)}</b></span>'
                f'<span><span style="color:#64748b;font-size:12px;font-weight:500">On Track</span> '
                f'<b style="color:#475569;font-size:16px;margin-left:4px">{len(on_track_lanes)}</b></span>'
                f'<span><span style="color:#64748b;font-size:12px;font-weight:500">At Risk</span> '
                f'<b style="color:#d97706;font-size:16px;margin-left:4px">{len(at_risk_lanes)}</b></span>'
                f'<span><span style="color:#64748b;font-size:12px;font-weight:500">Next Milestone</span> '
                f'<b style="color:#1d4ed8;font-size:14px;margin-left:4px">{next_ms_text}</b></span>'
                '</div>'
            )
            st.markdown(kpi_html, unsafe_allow_html=True)

            # ── vis-timeline Gantt (real swim lanes, sectioned + collapsible,
            # SVG overlay draws elbow dependency arrows) ────────────────────
            n_groups = (
                2  # section headers
                + len({r["project_id"] for r in filtered_rows if r["item_type"] == "dev_track"})
                + len({r["project_id"] for r in filtered_rows if r["item_type"] != "dev_track"})
            )
            # ~30 px per lane + ~80 px for axis/headers/margin.
            iframe_height = max(320, 90 + n_groups * 30)
            render_timeline(filtered_rows, today=today, height=iframe_height)

            # ── Phases needing attention ─────────────────────────────────────
            overdue = []
            upcoming = []
            for r in filtered_rows:
                if r["phase_status"] in ("Done", "Cancelled") or not r["end_date"]:
                    continue
                try:
                    end_dt = date.fromisoformat(r["end_date"])
                except ValueError:
                    continue
                lane_label = (r.get("track_name") or r.get("location")
                              or r["project_id"])
                if end_dt < today:
                    overdue.append((end_dt, lane_label, r["phase_name"]))
                elif (end_dt - today).days <= 30:
                    upcoming.append((end_dt, lane_label, r["phase_name"]))

            if overdue or upcoming:
                att_cols = st.columns(2)
                with att_cols[0]:
                    st.markdown("**Overdue phases**")
                    if overdue:
                        for end_dt, lane, phase in sorted(overdue):
                            d_over = (today - end_dt).days
                            st.markdown(
                                f"- {lane} — **{phase}** "
                                f"<span style='color:#b91c1c'>({d_over}d overdue)</span>",
                                unsafe_allow_html=True,
                            )
                    else:
                        st.caption("None.")
                with att_cols[1]:
                    st.markdown("**Ending within 30 days**")
                    if upcoming:
                        for end_dt, lane, phase in sorted(upcoming):
                            d_left = (end_dt - today).days
                            st.markdown(f"- {lane} — **{phase}** ({d_left}d)")
                    else:
                        st.caption("None.")


# ══════════════════════════════════════════════════════════════════════════════
# PROJECT DETAILS
# ══════════════════════════════════════════════════════════════════════════════

elif page == "Project Details":
    st.markdown("## Project Details")

    projects = get_all_projects()
    project_options = {f"{p['project_id']} — {p['location']} ({p['partner_org']})": p["project_id"] for p in projects}
    options_list = list(project_options.keys())

    # If navigated here via a Needs Attention click, pre-select that project.
    if "_pending_project_pid" in st.session_state:
        target_pid = st.session_state.pop("_pending_project_pid")
        for label, pid in project_options.items():
            if pid == target_pid:
                st.session_state["project_details_select"] = label
                break

    selected = st.selectbox(
        "Select a project", options_list, key="project_details_select"
    )

    if selected:
        pid = project_options[selected]
        p = get_project(pid)

        # ── Breadcrumb ────────────────────────────────────────────────────
        st.markdown(
            f'<div style="display:flex;align-items:center;gap:6px;font-size:12.5px;color:#616161;margin-bottom:20px">'
            f'<span style="color:#0078d4">Dashboard</span>'
            f'<span style="color:#c8c6c4">/</span>'
            f'<span style="color:#0078d4">Projects</span>'
            f'<span style="color:#c8c6c4">/</span>'
            f'<span>{p["project_id"]}</span></div>',
            unsafe_allow_html=True,
        )

        # ── Header Card with Metrics ─────────────────────────────────────
        target = p.get("target_date") or "TBD"
        target_sub = ""
        if p.get("target_date"):
            try:
                days_left = (date.fromisoformat(p["target_date"]) - date.today()).days
                target_sub = f"{days_left} days left" if days_left >= 0 else f"OVERDUE by {-days_left}d"
            except ValueError:
                pass
        cost = p.get("estimated_cost")
        cost_str = f"${cost:,.0f}" if cost else "TBD"
        last_updated = (p.get("last_updated") or "unknown")[:10]
        last_by = p.get("last_updated_by") or "unknown"
        owner = p.get("team_owner") or "Unassigned"
        target_sub_color = "#107c10" if "left" in target_sub else "#d13438" if "OVERDUE" in target_sub else "#616161"

        st.markdown(
            f'<div style="background:#fff;border:1px solid #edebe9;border-radius:8px;box-shadow:0 1px 2px rgba(0,0,0,0.04);padding:24px 28px;margin-bottom:20px;'
            f'box-shadow:0 1px 2px rgba(0,0,0,0.04)">'
            # Top: ID, Name, Status
            f'<div style="display:flex;align-items:flex-start;justify-content:space-between;flex-wrap:wrap;gap:12px">'
            f'<div>'
            f'<div style="display:flex;align-items:center;gap:14px;flex-wrap:wrap">'
            f'<span style="font-family:Cascadia Code,Consolas,monospace;font-size:13px;color:#0078d4;'
            f'background:#eff6fc;padding:3px 10px;border-radius:4px;font-weight:600;letter-spacing:0.5px">'
            f'{p["project_id"]}</span>'
            f'<span style="font-size:26px;font-weight:700;color:#242424;letter-spacing:-0.5px">'
            f'{p["location"]}, {p["country"]}</span>'
            f'{status_pill_html(p["status"])}'
            f'</div>'
            f'<div style="font-size:13.5px;color:#616161;margin-top:6px">'
            f'{p["partner_org"]} &middot; {p["country"]} &middot; {p["continent"]}</div>'
            f'</div></div>'
            # Metrics row
            f'<div style="display:grid;grid-template-columns:repeat(4,1fr);gap:0;margin-top:20px;'
            f'border-top:1px solid #edebe9;padding-top:20px">'
            # Owner
            f'<div style="padding:0 20px 0 0;border-right:1px solid #edebe9">'
            f'<div style="font-size:11.5px;color:#616161;text-transform:uppercase;letter-spacing:0.5px;'
            f'font-weight:600;margin-bottom:6px">Owner</div>'
            f'<div style="font-size:18px;font-weight:700;color:#242424;display:flex;align-items:center;gap:8px">'
            f'<span style="width:8px;height:8px;border-radius:50%;background:#0078d4"></span>{owner}</div>'
            f'<div style="font-size:12px;color:#616161;margin-top:3px">Project lead</div></div>'
            # Target Date
            f'<div style="padding:0 20px;border-right:1px solid #edebe9">'
            f'<div style="font-size:11.5px;color:#616161;text-transform:uppercase;letter-spacing:0.5px;'
            f'font-weight:600;margin-bottom:6px">Target Date</div>'
            f'<div style="font-size:18px;font-weight:700;color:#242424">{target}</div>'
            f'<div style="font-size:12px;color:{target_sub_color};margin-top:3px;font-weight:500">{target_sub}</div></div>'
            # Cost
            f'<div style="padding:0 20px;border-right:1px solid #edebe9">'
            f'<div style="font-size:11.5px;color:#616161;text-transform:uppercase;letter-spacing:0.5px;'
            f'font-weight:600;margin-bottom:6px">Estimated Cost</div>'
            f'<div style="font-size:18px;font-weight:700;color:#242424">{cost_str}</div>'
            f'<div style="font-size:12px;color:#616161;margin-top:3px">USD</div></div>'
            # Last Updated
            f'<div style="padding:0 0 0 20px">'
            f'<div style="font-size:11.5px;color:#616161;text-transform:uppercase;letter-spacing:0.5px;'
            f'font-weight:600;margin-bottom:6px">Last Updated</div>'
            f'<div style="font-size:18px;font-weight:700;color:#242424">{last_updated}</div>'
            f'<div style="font-size:12px;color:#616161;margin-top:3px">by {last_by}</div></div>'
            f'</div></div>',
            unsafe_allow_html=True,
        )

        # ── Edit Project ──────────────────────────────────────────────────
        with st.expander("✏️ Edit project"):
            with st.form(f"edit_project_form_{pid}", clear_on_submit=False):
                e1, e2 = st.columns(2)
                with e1:
                    ed_status = st.selectbox(
                        "Status", VALID_STATUSES,
                        index=VALID_STATUSES.index(p["status"]) if p.get("status") in VALID_STATUSES else 0,
                        key=f"ed_status_{pid}",
                    )
                    ed_owner = st.selectbox(
                        "Owner", [""] + TEAM_MEMBERS,
                        index=(TEAM_MEMBERS.index(p["team_owner"]) + 1) if p.get("team_owner") in TEAM_MEMBERS else 0,
                        key=f"ed_owner_{pid}",
                    )
                    ed_partner = st.text_input("Partner org", p.get("partner_org") or "", key=f"ed_partner_{pid}")
                    ed_deploy = st.text_input("Deployment type", p.get("deployment_type") or "", key=f"ed_deploy_{pid}")
                    ed_hardware = st.text_input("Hardware", p.get("hardware") or "", key=f"ed_hw_{pid}")
                with e2:
                    ed_timeline = st.text_input("Timeline label", p.get("timeline_label") or "", key=f"ed_tl_{pid}")
                    ed_target = st.text_input("Target date (YYYY-MM-DD)", p.get("target_date") or "", key=f"ed_td_{pid}")
                    confidence_options = ["", "hard", "committed", "soft", "aspirational"]
                    cur_conf = p.get("target_confidence") or ""
                    ed_conf = st.selectbox(
                        "Target confidence", confidence_options,
                        index=confidence_options.index(cur_conf) if cur_conf in confidence_options else 0,
                        key=f"ed_conf_{pid}",
                    )
                    ed_cost = st.number_input(
                        "Estimated cost (USD)",
                        value=float(p.get("estimated_cost") or 0.0),
                        min_value=0.0, step=1000.0, format="%.0f",
                        key=f"ed_cost_{pid}",
                    )
                ed_blocker = st.text_input("Blocker", p.get("blocker") or "", key=f"ed_blk_{pid}")
                ed_notes = st.text_area("Notes", p.get("notes") or "", height=100, key=f"ed_notes_{pid}")
                ed_by = st.selectbox("Updated by", TEAM_MEMBERS + ["Other"], key=f"ed_by_{pid}")

                save_clicked = st.form_submit_button(
                    "Save changes", type="primary",
                )

            if save_clicked:
                candidates = {
                    "status": ed_status,
                    "team_owner": ed_owner or None,
                    "partner_org": ed_partner or None,
                    "deployment_type": ed_deploy or None,
                    "hardware": ed_hardware or None,
                    "timeline_label": ed_timeline or None,
                    "target_date": ed_target or None,
                    "target_confidence": ed_conf or None,
                    "estimated_cost": ed_cost if ed_cost > 0 else None,
                    "blocker": ed_blocker or None,
                    "notes": ed_notes or None,
                }
                updates = {}
                for field, new_val in candidates.items():
                    old_val = p.get(field)
                    if str(old_val or "") != str(new_val or ""):
                        updates[field] = new_val

                if not updates:
                    st.info("No changes detected.")
                else:
                    try:
                        changes = update_project(pid, updates, ed_by)
                        if changes:
                            add_history(
                                pid, changes,
                                source_text="Project Details edit",
                                source_type="manual",
                                updated_by=ed_by,
                                llm_summary=f"Edited {', '.join(changes.keys())} via Project Details",
                            )
                        st.success(f"Saved {len(updates)} field(s).")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Save failed: {e}")

        # ── Phases editor ─────────────────────────────────────────────────
        with st.expander("🗂️ Phases (feeds the Timeline)"):
            from db import get_phases, upsert_phases
            from config import (
                DEV_PHASE_KEYS, DEPLOY_PHASE_KEYS, PHASE_STATUSES,
            )

            phase_key_options = (
                DEV_PHASE_KEYS if p.get("item_type") == "dev_track"
                else DEPLOY_PHASE_KEYS
            ) + ["custom"]

            current_phases = get_phases(pid)
            if current_phases:
                phases_df = pd.DataFrame(current_phases)
            else:
                phases_df = pd.DataFrame(columns=[
                    "id", "ordering", "phase_key", "name",
                    "start_date", "end_date", "status", "notes",
                ])

            # Normalize columns for the editor, preserving id as a hidden key.
            editor_cols = ["ordering", "phase_key", "name", "start_date",
                           "end_date", "status", "notes"]
            display_df = phases_df.copy()
            for col in editor_cols:
                if col not in display_df.columns:
                    display_df[col] = None
            # Keep id so we can match rows back on save.
            if "id" not in display_df.columns:
                display_df["id"] = None

            edited_phases = st.data_editor(
                display_df[["id"] + editor_cols],
                use_container_width=True,
                hide_index=True,
                num_rows="dynamic",
                key=f"phases_editor_{pid}",
                column_config={
                    "id": st.column_config.NumberColumn("id", disabled=True, help="phase row id"),
                    "ordering": st.column_config.NumberColumn("Order", width="small", default=0),
                    "phase_key": st.column_config.SelectboxColumn(
                        "Type", options=phase_key_options, default="custom", width="small",
                    ),
                    "name": st.column_config.TextColumn("Name", width="medium", required=True),
                    "start_date": st.column_config.TextColumn("Start (YYYY-MM-DD)", width="small"),
                    "end_date": st.column_config.TextColumn("End (YYYY-MM-DD)", width="small"),
                    "status": st.column_config.SelectboxColumn(
                        "Status", options=PHASE_STATUSES, default="Planned", width="small",
                    ),
                    "notes": st.column_config.TextColumn("Notes"),
                },
            )

            ph_cols = st.columns([5, 2])
            with ph_cols[1]:
                save_phases = st.button(
                    "Save phases", type="primary", use_container_width=True,
                    key=f"save_phases_{pid}",
                )

            if save_phases:
                rows_in = edited_phases.to_dict("records")
                # Pandas gives NaN for blank cells; coerce to None.
                for r in rows_in:
                    for k, v in list(r.items()):
                        if isinstance(v, float) and pd.isna(v):
                            r[k] = None
                        # Convert numpy ints to plain ints for sqlite
                        if hasattr(v, "item") and not isinstance(v, str):
                            try:
                                r[k] = v.item()
                            except (ValueError, AttributeError):
                                pass

                try:
                    deltas = upsert_phases(pid, rows_in)
                    n = (len(deltas["created"]) + len(deltas["updated"])
                         + len(deltas["deleted"]))
                    if n == 0:
                        st.info("No phase changes detected.")
                    else:
                        # One history row summarizing the phase edits.
                        summary_parts = []
                        if deltas["created"]:
                            summary_parts.append(f"{len(deltas['created'])} created")
                        if deltas["updated"]:
                            summary_parts.append(f"{len(deltas['updated'])} updated")
                        if deltas["deleted"]:
                            summary_parts.append(f"{len(deltas['deleted'])} deleted")
                        add_history(
                            pid,
                            {"phases": {"old": f"{len(current_phases)} phase(s)",
                                        "new": ", ".join(summary_parts)}},
                            source_text="Phase editor on Project Details",
                            source_type="manual",
                            updated_by="Phase editor",
                            llm_summary=f"Phases: {', '.join(summary_parts)}",
                        )
                        st.success(f"Phases saved — {', '.join(summary_parts)}.")
                        st.rerun()
                except Exception as e:
                    st.error(f"Save failed: {e}")

        # ── Two Column: Info + History / Contacts + Actions ───────────────
        left_col, right_col = st.columns([2, 1])

        with left_col:
            # Project Information (grid layout matching mockup)
            blocker = p.get("blocker") or "None"
            blocker_style = "color:#d13438;font-weight:500" if blocker != "None" else "color:#242424;font-weight:500"
            devops_html = ""
            if p.get("devops_id"):
                devops_html = (
                    f'<div style="grid-column:1/-1;padding:14px 0 14px 0">'
                    f'<div style="font-size:11.5px;color:#616161;text-transform:uppercase;letter-spacing:0.4px;'
                    f'font-weight:600;margin-bottom:5px">DevOps ID</div>'
                    f'<div style="font-size:13.5px;color:#242424;font-weight:500">'
                    f'<code style="font-family:Cascadia Code,Consolas,monospace;background:#f5f5f5;'
                    f'padding:1px 6px;border-radius:3px;font-size:12.5px">{p["devops_id"]}</code></div></div>'
                )
            st.markdown(
                f'<div style="background:#fff;border:1px solid #edebe9;border-radius:8px;box-shadow:0 1px 2px rgba(0,0,0,0.04);padding:24px;margin-bottom:20px;'
                f'box-shadow:0 1px 2px rgba(0,0,0,0.04)">'
                f'<div style="font-size:15px;font-weight:600;color:#242424;margin-bottom:16px">Project Information</div>'
                f'<div style="display:grid;grid-template-columns:1fr 1fr;gap:0">'
                # Row 1: Deployment Type | Hardware
                f'<div style="padding:14px 20px 14px 0;border-bottom:1px solid #f3f2f1">'
                f'<div style="font-size:11.5px;color:#616161;text-transform:uppercase;letter-spacing:0.4px;'
                f'font-weight:600;margin-bottom:5px">Deployment Type</div>'
                f'<div style="font-size:13.5px;color:#242424;font-weight:500">{p.get("deployment_type") or "TBD"}</div></div>'
                f'<div style="padding:14px 0 14px 20px;border-bottom:1px solid #f3f2f1;border-left:1px solid #f3f2f1">'
                f'<div style="font-size:11.5px;color:#616161;text-transform:uppercase;letter-spacing:0.4px;'
                f'font-weight:600;margin-bottom:5px">Hardware</div>'
                f'<div style="font-size:13.5px;color:#242424;font-weight:500">{p.get("hardware") or "TBD"}</div></div>'
                # Row 2: Timeline | Confidence
                f'<div style="padding:14px 20px 14px 0;border-bottom:1px solid #f3f2f1">'
                f'<div style="font-size:11.5px;color:#616161;text-transform:uppercase;letter-spacing:0.4px;'
                f'font-weight:600;margin-bottom:5px">Timeline</div>'
                f'<div style="font-size:13.5px;color:#242424;font-weight:500">{p.get("timeline_label") or "TBD"}</div></div>'
                f'<div style="padding:14px 0 14px 20px;border-bottom:1px solid #f3f2f1;border-left:1px solid #f3f2f1">'
                f'<div style="font-size:11.5px;color:#616161;text-transform:uppercase;letter-spacing:0.4px;'
                f'font-weight:600;margin-bottom:5px">Confidence</div>'
                f'<div style="font-size:13.5px;color:#242424;font-weight:500">{p.get("target_confidence") or "N/A"}</div></div>'
                # Full-width: Blocker
                f'<div style="grid-column:1/-1;padding:14px 0;border-bottom:1px solid #f3f2f1">'
                f'<div style="font-size:11.5px;color:#616161;text-transform:uppercase;letter-spacing:0.4px;'
                f'font-weight:600;margin-bottom:5px">Blocker</div>'
                f'<div style="font-size:13.5px;{blocker_style}">{blocker}</div></div>'
                # Full-width: Notes
                f'<div style="grid-column:1/-1;padding:14px 0;border-bottom:{"1px solid #f3f2f1" if p.get("devops_id") else "none"}">'
                f'<div style="font-size:11.5px;color:#616161;text-transform:uppercase;letter-spacing:0.4px;'
                f'font-weight:600;margin-bottom:5px">Notes</div>'
                f'<div style="font-size:13.5px;color:#424242;font-style:italic;font-weight:400;line-height:1.6">'
                f'{p.get("notes") or "None"}</div></div>'
                f'{devops_html}'
                f'</div></div>',
                unsafe_allow_html=True,
            )

            # ── History Timeline ───────────────────────────────────────────
            history = get_project_history(pid)

            # History header with filter
            hist_h_cols = st.columns([3, 2])
            with hist_h_cols[0]:
                st.markdown(
                    f'<div style="display:flex;align-items:center;gap:12px;margin-bottom:4px">'
                    f'<span style="font-size:17px;font-weight:700;color:#242424;letter-spacing:-0.2px">History</span>'
                    f'<span style="font-size:12px;color:#616161">({len(history) if history else 0} entries)</span>'
                    f'</div>',
                    unsafe_allow_html=True,
                )
            with hist_h_cols[1]:
                source_filter = st.selectbox(
                    "Filter",
                    ["All", "Email", "Manual", "System", "Teams"],
                    key="history_filter",
                    label_visibility="collapsed",
                )

            if history:
                filtered_history = history
                if source_filter != "All":
                    filter_map = {"Email": "email", "Manual": "manual_note", "System": "system", "Teams": "teams_paste"}
                    filtered_history = [h for h in history if h.get("source_type") == filter_map.get(source_filter, "")]

                # Build timeline in a single HTML block for the vertical line effect
                timeline_html = '<div class="sparrow-timeline">'
                for i, h in enumerate(filtered_history[:20]):
                    expanded = i < 3
                    timeline_html += timeline_entry_html(
                        timestamp=h.get("timestamp", ""),
                        source_type=h.get("source_type", "manual"),
                        author=h.get("updated_by", "unknown"),
                        summary=h.get("llm_summary", "Update"),
                        changes=h.get("changes") if expanded else None,
                        source_text=h.get("source_text") if expanded else None,
                        expanded=expanded,
                    )
                timeline_html += '</div>'
                st.markdown(timeline_html, unsafe_allow_html=True)

                if len(filtered_history) > 20:
                    st.caption(f"Showing 20 of {len(filtered_history)} entries")
            else:
                st.info("No history yet.")

        with right_col:
            # Contacts
            project_contacts = get_contacts(pid)
            st.markdown(
                '<div style="background:#fff;border:1px solid #edebe9;border-radius:8px;box-shadow:0 1px 2px rgba(0,0,0,0.04);'
                'padding:16px;margin-bottom:12px">'
                '<div style="font-size:14px;font-weight:600;margin-bottom:10px">Contacts</div>',
                unsafe_allow_html=True,
            )
            if project_contacts:
                for c in project_contacts:
                    st.markdown(
                        f'<div style="padding:6px 0;border-bottom:1px solid #f3f2f1;font-size:13px">'
                        f'<strong>{c.get("name", "")}</strong>'
                        f'<div style="color:#616161">{c.get("organization", "")} · {c.get("role", "")}</div>'
                        f'{"<div style=color:#0078d4>" + c["email"] + "</div>" if c.get("email") else ""}'
                        f'</div>',
                        unsafe_allow_html=True,
                    )
            else:
                st.markdown('<div style="color:#8a8886;font-size:13px">No contacts yet</div>',
                            unsafe_allow_html=True)
            st.markdown('</div>', unsafe_allow_html=True)

            # Active Alerts
            project_nudges = get_active_nudges(pid)
            if project_nudges:
                st.markdown(
                    '<div style="background:#fff;border:1px solid #edebe9;border-radius:8px;box-shadow:0 1px 2px rgba(0,0,0,0.04);'
                    'padding:16px;margin-bottom:12px">'
                    '<div style="font-size:14px;font-weight:600;margin-bottom:10px">Active Alerts</div>',
                    unsafe_allow_html=True,
                )
                for n in project_nudges:
                    sev_color = {"info": COLORS["primary"], "warning": COLORS["warning"],
                                 "escalation": COLORS["danger"]}.get(n["severity"], COLORS["neutral"])
                    st.markdown(
                        f'<div style="border-left:3px solid {sev_color};padding:8px 12px;margin-bottom:8px;'
                        f'font-size:13px">{severity_badge_html(n["severity"])} {n["message"][:150]}</div>',
                        unsafe_allow_html=True,
                    )
                st.markdown('</div>', unsafe_allow_html=True)

            # Quick Actions
            st.markdown(
                '<div style="background:#fff;border:1px solid #edebe9;border-radius:8px;box-shadow:0 1px 2px rgba(0,0,0,0.04);'
                'padding:16px">'
                '<div style="font-size:14px;font-weight:600;margin-bottom:10px">Quick Actions</div>',
                unsafe_allow_html=True,
            )
            st.button("Submit update for this project", use_container_width=True, key="quick_update")
            st.markdown(
                f'<div style="padding:6px 0;font-size:13px;color:#0078d4;cursor:pointer">'
                f'Export history</div></div>',
                unsafe_allow_html=True,
            )


# ══════════════════════════════════════════════════════════════════════════════
# SPRINTS
# ══════════════════════════════════════════════════════════════════════════════

elif page == "Sprints":
    from devops_sync import (
        sync_all, get_work_items, get_work_items_by_sprint,
        get_work_items_by_person, get_last_sync_time,
    )
    from config import AZURE_DEVOPS_PAT, AZURE_DEVOPS_ORG, AZURE_DEVOPS_PROJECT

    st.markdown("## Sprints & Roadmap")

    devops_configured = bool(AZURE_DEVOPS_PAT)

    # ── Sync Controls ────────────────────────────────────────────────────
    sync_cols = st.columns([4, 1, 1])
    with sync_cols[0]:
        last_sync = get_last_sync_time()
        if last_sync:
            st.markdown(
                f'<div style="font-size:13px;color:#616161;padding-top:8px">'
                f'Last synced: <strong>{last_sync}</strong> · '
                f'{AZURE_DEVOPS_ORG}/{AZURE_DEVOPS_PROJECT}</div>',
                unsafe_allow_html=True,
            )
        else:
            st.markdown(
                '<div style="font-size:13px;color:#8a8886;padding-top:8px">'
                'Not synced yet — click Sync to pull from Azure DevOps</div>',
                unsafe_allow_html=True,
            )
    with sync_cols[2]:
        if st.button("🔄 Sync", type="primary", disabled=not devops_configured,
                      use_container_width=True):
            with st.spinner("Syncing from Azure DevOps..."):
                try:
                    result = sync_all()
                    st.success(f"Synced {result['iterations']} sprints, {result['work_items']} work items")
                    st.rerun()
                except Exception as e:
                    st.error(f"Sync failed: {e}")

    if not devops_configured:
        st.warning("Azure DevOps PAT not configured. Add `AZURE_DEVOPS_PAT` to your `.env` file.")

    # ── Shared helpers for both tabs ─────────────────────────────────────
    state_colors = {
        "Done": COLORS["success"], "Closed": COLORS["success"],
        "Active": COLORS["primary"], "In Progress": COLORS["primary"],
        "New": COLORS["neutral"], "To Do": COLORS["neutral"],
        "Resolved": "#5c2d91",
    }
    type_icons = {
        "User Story": "📖", "Task": "✅", "Bug": "🐛",
        "Feature": "⭐", "Epic": "🏔️", "Issue": "⚠️",
    }

    def _sprint_sort_key(name: str):
        # "April 2026" → parseable → sort by that date descending
        # unparseable names (e.g. "Dec-Jan-25-26") → push below parsed ones
        # "No sprint assigned" → push to the very bottom
        from devops_sync import NO_SPRINT_LABEL
        if name == NO_SPRINT_LABEL:
            return (0, date.min)
        try:
            d = datetime.strptime(name, "%B %Y").date()
            return (2, d)
        except ValueError:
            return (1, date.min)

    def _one_thing_tag_for(sprint_name: str) -> str:
        # Tag convention: "One Thing - <Month> FY<YY>" (e.g. "One Thing - April FY26").
        try:
            sprint_date = datetime.strptime(sprint_name, "%B %Y").date()
            return f"One Thing - {sprint_date.strftime('%B')} FY{sprint_date.year % 100:02d}"
        except ValueError:
            return ""

    # ── View Tabs ────────────────────────────────────────────────────────
    tab_board, tab_person = st.tabs(["Sprint Board", "By Person"])

    # ── Sprint Board ─────────────────────────────────────────────────────
    with tab_board:
        sprints_data = get_work_items_by_sprint()

        if not sprints_data:
            st.info("No work items found. Sync from Azure DevOps to populate.")
        else:
            for sprint_name in sorted(sprints_data.keys(), key=_sprint_sort_key, reverse=True):
                items = sprints_data[sprint_name]
                total = len(items)

                st.markdown(
                    f'<div style="background:#fff;border:1px solid #edebe9;border-radius:8px;box-shadow:0 1px 2px rgba(0,0,0,0.04);'
                    f'padding:18px 22px;margin-bottom:16px">'
                    f'<div style="display:flex;align-items:center;justify-content:space-between;'
                    f'flex-wrap:wrap;gap:8px;margin-bottom:14px">'
                    f'<div>'
                    f'<span style="font-size:16px;font-weight:700;color:#242424">{sprint_name}</span>'
                    f'</div>'
                    f'<div style="display:flex;align-items:center;gap:12px">'
                    f'<span style="font-size:12px;color:#616161">{total} open</span>'
                    f'</div></div>',
                    unsafe_allow_html=True,
                )

                # Items tagged as the sprint's "One Thing" go on top.
                one_thing_tag = _one_thing_tag_for(sprint_name)

                def _item_sort(wi):
                    is_one_thing = bool(one_thing_tag) and one_thing_tag in (wi.get("tags") or "")
                    return (0 if is_one_thing else 1, wi.get("state", ""), wi.get("title", ""))

                # Work items table
                items_html = ""
                for wi in sorted(items, key=_item_sort):
                    sc = state_colors.get(wi.get("state", ""), COLORS["neutral"])
                    icon = type_icons.get(wi.get("work_item_type", ""), "📋")
                    person = wi.get("assigned_to") or "Unassigned"
                    person_short = person.split(" ")[0] if person != "Unassigned" else person
                    tags = wi.get("tags") or ""
                    is_one_thing = bool(one_thing_tag) and one_thing_tag in tags
                    tags_html = ""
                    if tags:
                        for tag in tags.split(";")[:3]:
                            tag = tag.strip()
                            if tag:
                                tags_html += (
                                    f'<span style="display:inline-block;padding:1px 6px;'
                                    f'background:#f3f2f1;border-radius:3px;font-size:10px;'
                                    f'color:#616161;margin-left:4px">{tag}</span>'
                                )
                    row_bg = "#fff8e1" if is_one_thing else "transparent"
                    star = '<span title="One Thing" style="color:#f2c811">★</span>' if is_one_thing else '<span style="width:14px"></span>'
                    items_html += (
                        f'<div style="display:flex;align-items:center;gap:10px;'
                        f'padding:8px 6px;border-bottom:1px solid #f3f2f1;font-size:13px;'
                        f'background:{row_bg}">'
                        f'{star}'
                        f'<span style="width:8px;height:8px;border-radius:50%;background:{sc};flex-shrink:0"></span>'
                        f'<span>{icon}</span>'
                        f'<span style="flex:1;font-weight:500;color:#242424">{wi["title"]}</span>'
                        f'{tags_html}'
                        f'<span style="font-size:12px;color:#616161;min-width:80px;text-align:right">{person_short}</span>'
                        f'<span style="display:inline-block;padding:2px 8px;border-radius:10px;'
                        f'font-size:11px;font-weight:600;background:{sc}18;color:{sc}">{wi.get("state", "?")}</span>'
                        f'</div>'
                    )
                st.markdown(items_html + '</div>', unsafe_allow_html=True)

    # ── By Person ────────────────────────────────────────────────────────
    with tab_person:
        person_data = get_work_items_by_person()
        if not person_data:
            st.info("No work items found.")
        else:
            from devops_sync import NO_SPRINT_LABEL

            def _sprint_of(wi):
                path = wi.get("iteration_path") or ""
                return path.split("\\")[-1] if path else NO_SPRINT_LABEL

            for person in sorted(person_data.keys()):
                items = person_data[person]
                active = sum(1 for i in items if i.get("state") not in ("Done", "Closed", "Removed"))
                done = sum(1 for i in items if i.get("state") in ("Done", "Closed"))

                st.markdown(
                    f'<div style="background:#fff;border:1px solid #edebe9;border-radius:8px;box-shadow:0 1px 2px rgba(0,0,0,0.04);'
                    f'padding:16px 20px;margin-bottom:12px">'
                    f'<div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:10px">'
                    f'<div style="display:flex;align-items:center;gap:10px">'
                    f'<div style="width:32px;height:32px;border-radius:50%;background:#0078d4;'
                    f'display:flex;align-items:center;justify-content:center;color:#fff;'
                    f'font-size:14px;font-weight:600">{person[0].upper() if person else "?"}</div>'
                    f'<span style="font-size:15px;font-weight:600;color:#242424">{person}</span>'
                    f'</div>'
                    f'<div style="font-size:12px;color:#616161">'
                    f'{active} active · {done} done</div>'
                    f'</div>',
                    unsafe_allow_html=True,
                )

                grouped = {}
                for wi in items:
                    grouped.setdefault(_sprint_of(wi), []).append(wi)

                body_html = ""
                for sprint_name in sorted(grouped.keys(), key=_sprint_sort_key, reverse=True):
                    one_thing_tag = _one_thing_tag_for(sprint_name)

                    def _item_sort(wi, tag=one_thing_tag):
                        is_one_thing = bool(tag) and tag in (wi.get("tags") or "")
                        return (0 if is_one_thing else 1, wi.get("state", ""), wi.get("title", ""))

                    body_html += (
                        f'<div style="font-size:12px;font-weight:600;color:#616161;'
                        f'text-transform:uppercase;letter-spacing:0.5px;margin:10px 0 4px">'
                        f'{sprint_name}</div>'
                    )
                    for wi in sorted(grouped[sprint_name], key=_item_sort):
                        sc = state_colors.get(wi.get("state", ""), COLORS["neutral"])
                        is_one_thing = bool(one_thing_tag) and one_thing_tag in (wi.get("tags") or "")
                        row_bg = "#fff8e1" if is_one_thing else "transparent"
                        star = ('<span title="One Thing" style="color:#f2c811">★</span>'
                                if is_one_thing else '<span style="width:14px"></span>')
                        body_html += (
                            f'<div style="display:flex;align-items:center;gap:8px;'
                            f'padding:6px 6px;border-bottom:1px solid #f3f2f1;font-size:13px;'
                            f'background:{row_bg}">'
                            f'{star}'
                            f'<span style="width:6px;height:6px;border-radius:50%;background:{sc};flex-shrink:0"></span>'
                            f'<span style="flex:1;color:#242424">{wi["title"]}</span>'
                            f'<span style="display:inline-block;padding:2px 8px;border-radius:10px;'
                            f'font-size:11px;font-weight:600;background:{sc}18;color:{sc}">{wi.get("state", "?")}</span>'
                            f'</div>'
                        )

                st.markdown(body_html + '</div>', unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# REPORTS
# ══════════════════════════════════════════════════════════════════════════════

elif page == "Reports":
    st.markdown("## Reports")
    st.markdown("Generate insights and export data.")

    tab_quick, tab_custom, tab_export = st.tabs(["Quick Reports", "Custom Report", "Export"])

    with tab_quick:
        report_options = [
            ("📊", "Executive Summary", "High-level overview for leadership"),
            ("🌍", "Regional Breakdown", "Projects grouped by continent and country"),
            ("🚨", "Blocked & At Risk", "All projects with blockers or risk flags"),
            ("📅", "Timeline Status", "Upcoming deadlines and overdue projects"),
            ("🤖", "Robin Dependencies", "All projects waiting on ROBIN hardware"),
            ("💰", "Cost Summary", "Budget overview across all installations"),
        ]

        # 2x3 grid
        for row in range(2):
            cols = st.columns(3)
            for col_idx in range(3):
                idx = row * 3 + col_idx
                icon, title, desc = report_options[idx]
                with cols[col_idx]:
                    st.markdown(
                        f'<div style="background:#fff;border:1px solid #edebe9;border-radius:8px;box-shadow:0 1px 2px rgba(0,0,0,0.04);'
                        f'padding:20px;text-align:center;min-height:140px">'
                        f'<div style="font-size:28px;margin-bottom:8px">{icon}</div>'
                        f'<div style="font-size:14px;font-weight:600">{title}</div>'
                        f'<div style="font-size:12px;color:#616161;margin-top:4px">{desc}</div>'
                        f'</div>',
                        unsafe_allow_html=True,
                    )
                    if st.button(f"Generate", key=f"report_{idx}", use_container_width=True,
                                 disabled=not llm_available):
                        st.session_state["report_type"] = title

        days_range = st.slider("Include history from last N days", 7, 90, 30)

        if st.session_state.get("report_type"):
            with st.spinner(f"Generating {st.session_state['report_type']}..."):
                from llm import generate_report
                report = generate_report(st.session_state["report_type"], days=days_range)
            st.markdown("---")
            st.markdown(report)
            del st.session_state["report_type"]

    with tab_custom:
        st.markdown("Describe the report you want in plain English:")
        custom_request = st.text_area(
            "Report request", height=100,
            placeholder="e.g., Show me all South America projects grouped by country with their current blockers",
        )
        if st.button("Generate Custom Report", type="primary", disabled=not llm_available or not custom_request.strip()):
            with st.spinner("Generating..."):
                from llm import generate_report
                report = generate_report(custom_request)
            st.markdown("---")
            st.markdown(report)

    with tab_export:
        export_cols = st.columns(2)
        with export_cols[0]:
            st.markdown(
                '<div style="background:#fff;border:1px solid #edebe9;border-radius:8px;box-shadow:0 1px 2px rgba(0,0,0,0.04);padding:20px">'
                '<div style="font-size:14px;font-weight:600">Projects CSV</div>'
                '<div style="font-size:13px;color:#616161;margin:8px 0">Current state of all projects</div>'
                '</div>',
                unsafe_allow_html=True,
            )
            projects = get_all_projects()
            df = pd.DataFrame(projects)
            csv = df.to_csv(index=False)
            st.download_button("Download Projects", csv, "sparrow_projects.csv", "text/csv",
                               use_container_width=True)

        with export_cols[1]:
            st.markdown(
                '<div style="background:#fff;border:1px solid #edebe9;border-radius:8px;box-shadow:0 1px 2px rgba(0,0,0,0.04);padding:20px">'
                '<div style="font-size:14px;font-weight:600">History CSV</div>'
                '<div style="font-size:13px;color:#616161;margin:8px 0">Full changelog of all updates</div>'
                '</div>',
                unsafe_allow_html=True,
            )
            all_history = get_recent_history(days=365, limit=5000)
            if all_history:
                for h in all_history:
                    h["changes"] = json.dumps(h["changes"])
                hdf = pd.DataFrame(all_history)
                hcsv = hdf.to_csv(index=False)
                st.download_button("Download History", hcsv, "sparrow_history.csv", "text/csv",
                                   use_container_width=True)


# ══════════════════════════════════════════════════════════════════════════════
# SETTINGS
# ══════════════════════════════════════════════════════════════════════════════

elif page == "Settings":
    st.markdown("## Settings")
    st.markdown("Configure monitoring, team, and integrations.")

    tab_team, tab_monitor, tab_devops, tab_email, tab_notif, tab_ai = st.tabs(
        ["Team", "Monitor", "Azure DevOps", "Email Ingestion", "Notifications", "AI / LLM"]
    )

    with tab_team:
        st.markdown("### Team Members")
        team_data = []
        for member in TEAM_MEMBERS:
            team_data.append({"Name": member, "Email": "Not configured", "Role": "Member"})
        st.dataframe(pd.DataFrame(team_data), use_container_width=True, hide_index=True)
        st.caption("Edit team members in config.py → TEAM_MEMBERS")

    with tab_monitor:
        st.markdown("### Staleness Thresholds")
        st.markdown("Projects are flagged as stale when they haven't been updated within these windows:")
        threshold_df = pd.DataFrame(
            [{"Status": k, "Stale After (days)": v} for k, v in STALENESS_THRESHOLDS.items()]
        )
        st.dataframe(threshold_df, use_container_width=True, hide_index=True)

        st.markdown("### Deadline Alert Windows")
        from config import DEADLINE_ALERTS
        for alert in DEADLINE_ALERTS:
            days = alert["days_before"]
            sev = alert["severity"]
            label = "Overdue" if days == 0 else f"{days} days before"
            st.markdown(f"- {severity_badge_html(sev)} — {label}", unsafe_allow_html=True)

        st.markdown("### Run Monitor")
        use_llm = st.checkbox("Use LLM for nudge generation", value=llm_available, disabled=not llm_available)
        if st.button("Run Monitor Check", type="primary"):
            with st.spinner("Checking all projects..."):
                from monitor import run_monitor
                import io
                from contextlib import redirect_stdout

                f = io.StringIO()
                with redirect_stdout(f):
                    nudges = run_monitor(use_llm=use_llm, dry_run=False, send_email=False)
                output = f.getvalue()

            if nudges:
                st.warning(f"Found {len(nudges)} project(s) needing attention.")
                for n in nudges:
                    st.markdown(f"**{n['project_id']}** — {n['type']} ({n['severity']})")
                    st.markdown(n["message"][:300])
                    st.markdown("---")
            else:
                st.success("All projects are up to date.")
            with st.expander("Raw output"):
                st.text(output)

        st.markdown("### Schedule (Cron)")
        st.code(
            "# Add to crontab (crontab -e) — runs weekdays at 9am:\n"
            "0 9 * * 1-5  cd /path/to/sparrow-tracker && python3 monitor.py --send-email",
            language="bash",
        )

    with tab_devops:
        st.markdown("### Azure DevOps Integration")
        from config import AZURE_DEVOPS_PAT, AZURE_DEVOPS_ORG, AZURE_DEVOPS_PROJECT, DEVOPS_SEARCH_TERMS
        devops_configured = bool(AZURE_DEVOPS_PAT)
        if devops_configured:
            st.success(f"Connected to Azure DevOps")
            st.markdown(f"**Organization:** `{AZURE_DEVOPS_ORG}`")
            st.markdown(f"**Project:** `{AZURE_DEVOPS_PROJECT}`")
            st.markdown(f"**Search terms:** {', '.join(DEVOPS_SEARCH_TERMS)}")

            from devops_sync import get_last_sync_time, sync_all
            last_sync = get_last_sync_time()
            if last_sync:
                st.markdown(f"**Last synced:** {last_sync}")

            if st.button("Sync Now", type="primary", key="settings_devops_sync"):
                with st.spinner("Syncing from Azure DevOps..."):
                    try:
                        result = sync_all()
                        st.success(
                            f"Synced {result['work_items']} work items "
                            f"({result['sprint_query']} in current sprint query)"
                        )
                    except Exception as e:
                        st.error(f"Sync failed: {e}")

            if st.button("Test Connection", key="test_devops"):
                try:
                    from devops_sync import fetch_work_item_ids
                    ids = fetch_work_item_ids()
                    st.success(f"Connection OK — {len(ids)} matching work items")
                except Exception as e:
                    st.error(f"Connection failed: {e}")
        else:
            st.warning(
                "Not configured. Either run `az login` or add `AZURE_DEVOPS_PAT` to `.env`."
            )

        st.code(
            "# Preferred: run `az login` — auth uses Entra ID.\n"
            "# Fallback (PAT rotates weekly under Microsoft policy):\n"
            "AZURE_DEVOPS_PAT=your-personal-access-token\n"
            "AZURE_DEVOPS_ORG=onecela\n"
            "AZURE_DEVOPS_PROJECT=AI For Good Lab",
            language="bash",
        )

    with tab_email:
        st.markdown("### Email Ingestion (IMAP)")
        st.markdown("Configure a mailbox for automatic email processing.")
        imap_configured = bool(IMAP_HOST)
        if imap_configured:
            st.success(f"IMAP configured: {IMAP_HOST}")
        else:
            st.info("Not configured. Add IMAP settings to your `.env` file.")

        st.code(
            "# Add to .env:\n"
            "IMAP_HOST=imap.example.com\n"
            "IMAP_USER=sparrow-inbox@example.com\n"
            "IMAP_PASS=your-password\n"
            "IMAP_FOLDER=INBOX\n"
            "IMAP_DONE_FOLDER=Processed",
            language="bash",
        )

        if imap_configured:
            if st.button("Test Connection"):
                try:
                    from email_ingest import connect_imap
                    imap = connect_imap()
                    imap.logout()
                    st.success("Connection successful!")
                except Exception as e:
                    st.error(f"Connection failed: {e}")

    with tab_notif:
        st.markdown("### Notification Settings (SMTP)")
        from config import SMTP_HOST
        if SMTP_HOST:
            st.success(f"SMTP configured: {SMTP_HOST}")
        else:
            st.info("Not configured. Add SMTP settings to your `.env` file.")

        st.code(
            "# Add to .env:\n"
            "SPARROW_SMTP_HOST=smtp.example.com\n"
            "SPARROW_SMTP_PORT=587\n"
            "SPARROW_SMTP_USER=your-user\n"
            "SPARROW_SMTP_PASS=your-password\n"
            "SPARROW_NOTIFY_FROM=sparrow-tracker@noreply.local",
            language="bash",
        )

    with tab_ai:
        st.markdown("### AI / LLM Configuration")
        if llm_available:
            st.success(f"Connected to Azure OpenAI")
            st.markdown(f"**Endpoint:** `{AZURE_OPENAI_ENDPOINT[:40]}...`")
            st.markdown(f"**Deployment:** `{AZURE_OPENAI_DEPLOYMENT}`")

            if st.button("Test Connection", key="test_llm"):
                try:
                    from llm import _chat
                    result = _chat("You are a test.", "Say hello in one word.")
                    st.success(f"Connected — Response: \"{result}\"")
                except Exception as e:
                    st.error(f"Connection failed: {e}")
        else:
            st.warning("Not configured. Add Azure OpenAI settings to your `.env` file.")

        st.code(
            "# Add to .env:\n"
            "AZURE_OPENAI_ENDPOINT=https://your-resource.openai.azure.com/\n"
            "AZURE_OPENAI_DEPLOYMENT=gpt-4\n"
            "AZURE_OPENAI_API_KEY=your-key\n"
            "AZURE_OPENAI_API_VERSION=2024-10-21",
            language="bash",
        )
