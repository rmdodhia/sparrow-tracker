#!/usr/bin/env python3
"""
One-time import: read SPARROW Installation Requests from Excel and seed the SQLite database.
Run: python3 seed_data.py
"""

import os
import sys
import openpyxl
from datetime import date

# Ensure we can import sibling modules
sys.path.insert(0, os.path.dirname(__file__))

from db import init_db, create_project, add_history
from config import FY26_END


# ── Project ID generation ─────────────────────────────────────────────────────

CONTINENT_CODES = {
    "Africa": "AF", "Antarctica": "AN", "Asia": "AS",
    "Europe": "EU", "North America": "NA", "Oceania": "OC",
    "South America": "SA",
}

COUNTRY_CODES = {
    "DRC": "DRC", "Kenya": "KEN", "Tanzania": "TZA", "UK": "UK",
    "Indonesia ": "IDN", "Indonesia": "IDN", "Japan": "JPN",
    "Greenland": "GRL", "Scotland": "SCO", "USA": "USA",
    "New Zeland": "NZL", "Brazil": "BRA", "Colombia": "COL",
    "Peru": "PER", "Uruguay": "URY", "Antarctica": "ANT",
    "Chile": "CHL", "Atlantic Ocean": "ATL",
}


def make_project_id(continent, country, location, seen: set) -> str:
    cc = CONTINENT_CODES.get(continent, "XX")
    ctry = COUNTRY_CODES.get(country.strip(), country.strip()[:3].upper())
    loc_chars = "".join(c for c in location if c.isalpha())[:4].upper()
    base = f"{cc}-{ctry}-{loc_chars}"
    if base not in seen:
        seen.add(base)
        return base
    for i in range(2, 10):
        candidate = f"{base}{i}"
        if candidate not in seen:
            seen.add(candidate)
            return candidate
    return base


# ── Status normalization ──────────────────────────────────────────────────────

def normalize_status(raw_status, notes=""):
    """Map the messy Excel STATUS to a clean enum value. Returns (status, extra_notes)."""
    if not raw_status:
        return "Scoping", ""

    s = str(raw_status).strip()
    sl = s.lower()

    if "complete" in sl or sl == "done":
        return "Complete", ""
    if "installed" in sl:
        return "Installed", ""
    if "active" in sl:
        return "Active", ""
    if "approved" in sl:
        return "Approved", ""
    if "on hold" in sl:
        return "On Hold", s  # keep the reason
    if "blocked" in sl:
        return "Blocked", s
    if "at risk" in sl or "descoped" in sl:
        return "At Risk", s
    if "waiting" in sl:
        return "Waiting", s
    if "on track" in sl:
        return "Active", s if len(s) > 20 else ""  # long text = extra notes
    if "scoping" in sl or "early scoping" in sl:
        return "Scoping", ""
    # If status looks like it has narrative text (>50 chars), it's probably misplaced data
    if len(s) > 50:
        return "Scoping", s

    return "Scoping", ""


# ── Timeline parsing ──────────────────────────────────────────────────────────

def parse_timeline(raw_timeline):
    """Returns (timeline_label, target_date_iso, confidence)."""
    if not raw_timeline:
        return "TBD", None, None

    s = str(raw_timeline).strip()
    sl = s.lower()

    if sl in ("done", "tbd", ""):
        return s, None, None

    # Excel date serial number (e.g., 45962 for Ignite)
    if s.isdigit() and int(s) > 40000:
        return "Done", None, None  # Ignite was a past event

    # "Before End FY26?" or "Before End FY26"
    aspirational = "?" in s
    if "fy26" in sl or "fy 26" in sl:
        confidence = "aspirational" if aspirational else "soft"
        if "may" in sl:
            return s, "2026-05-31", confidence
        return s, str(FY26_END), confidence

    # "Target Dec 2026", "Target Jun 2026", "Target Mar 2026"
    month_map = {
        "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
        "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
    }
    for m_name, m_num in month_map.items():
        if m_name in sl:
            # Find a year
            for y in range(2025, 2029):
                if str(y) in s:
                    import calendar
                    last_day = calendar.monthrange(y, m_num)[1]
                    return s, f"{y}-{m_num:02d}-{last_day:02d}", "soft"

    # "Q4 2026"
    if s.lower().startswith("q"):
        for y in range(2025, 2029):
            if str(y) in s:
                q = int(s[1])
                end_month = q * 3
                import calendar
                last_day = calendar.monthrange(y, end_month)[1]
                return s, f"{y}-{end_month:02d}-{last_day:02d}", "soft"

    # "2026-2027"
    if "-" in s and all(p.strip().isdigit() for p in s.split("-")):
        parts = s.split("-")
        if len(parts) == 2 and int(parts[1]) > 2000:
            return s, f"{parts[1].strip()}-06-30", "aspirational"

    return s, None, None


# ── Main seed logic ───────────────────────────────────────────────────────────

# Manually fixed data for rows where columns are misaligned in the Excel
ROW_OVERRIDES = {
    # Row 9: Antarctica Chile - data is scattered across wrong columns
    9: {
        "status": "On Hold",
        "deployment_type": "TBD",
        "timeline_raw": "TBD",
        "hardware": "TBD",
        "cost": None,
        "notes": "Winter Sparrow work paused until Water Sparrow is proven; thermal and battery issues remain.",
    },
    # Row 10: Antarctica Uruguay Base - same issue
    10: {
        "status": "Scoping",
        "deployment_type": "TBD",
        "timeline_raw": "TBD",
        "hardware": "TBD",
        "cost": None,
        "notes": "Bruno plans to speak to Air Force contacts; no partner currently in place. Early scoping, no partner yet.",
    },
    # Row 26: Mt Rainier - status column has narrative text
    26: {
        "status": "Scoping",
        "notes": "Water-level monitoring may not need a full Sparrow unit; Bruno is reviewing the one-pager.",
    },
}


def seed():
    xlsx_path = os.path.join(os.path.dirname(__file__), "..", "SPARROW_BACKLOG_OF_PRIORITIES.xlsx")
    if not os.path.exists(xlsx_path):
        print(f"ERROR: Excel file not found at {xlsx_path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["SPARROW Installation Requests"]

    init_db()
    seen_ids = set()
    seeded = 0

    for row_num in range(2, ws.max_row + 1):  # skip header
        vals = [ws.cell(row_num, c).value for c in range(1, ws.max_column + 1)]
        if not any(vals):
            continue

        # Unpack columns per header:
        # Continent, Country/Region, Location, Partner Organizations, STATUS,
        # Deployment Type, Timeline, Hardware Details, Estimated Total Cost (USD),
        # Notes, DevOps ID, Carl comments, Miao Comments
        continent    = vals[0] or ""
        country      = vals[1] or ""
        location     = vals[2] or ""
        partner_org  = vals[3] or ""
        raw_status   = vals[4]
        deploy_type  = vals[5] or ""
        raw_timeline = vals[6]
        hardware     = vals[7] or ""
        raw_cost     = vals[8]
        notes        = vals[9] or ""
        devops_id    = vals[10]
        carl_comment = vals[11] or ""
        miao_comment = vals[12] or ""

        # Apply manual overrides for broken rows
        override = ROW_OVERRIDES.get(row_num, {})
        if override:
            raw_status   = override.get("status", raw_status)
            deploy_type  = override.get("deployment_type", deploy_type)
            raw_timeline = override.get("timeline_raw", raw_timeline)
            hardware     = override.get("hardware", hardware)
            if "cost" in override:
                raw_cost = override["cost"]
            if "notes" in override:
                notes = override["notes"]

        # Normalize
        status, extra_notes = normalize_status(raw_status, notes)
        if extra_notes and extra_notes not in (notes or ""):
            notes = f"{extra_notes}; {notes}" if notes else extra_notes

        # Merge Carl/Miao comments into notes
        if carl_comment:
            notes = f"{notes}; [Carl] {carl_comment}" if notes else f"[Carl] {carl_comment}"
        if miao_comment:
            notes = f"{notes}; [Miao] {miao_comment}" if notes else f"[Miao] {miao_comment}"

        timeline_label, target_date, target_confidence = parse_timeline(raw_timeline)

        # Cost
        cost = None
        if raw_cost is not None and str(raw_cost).strip().replace(".", "").isdigit():
            cost = float(raw_cost)

        # DevOps ID
        devops = int(devops_id) if devops_id and str(devops_id).strip().isdigit() else None

        # Generate project ID
        pid = make_project_id(continent, country, location, seen_ids)

        # Guess team owner from notes context (basic heuristic)
        owner = None
        notes_lower = (notes or "").lower() + " " + (str(raw_status) or "").lower()
        if "bruno" in notes_lower:
            owner = "Bruno"
        elif "carl" in notes_lower:
            owner = "Carl"
        elif "miao" in notes_lower:
            owner = "Miao"

        project_data = {
            "project_id":        pid,
            "continent":         continent.strip(),
            "country":           country.strip(),
            "location":          location.strip(),
            "partner_org":       partner_org.strip(),
            "status":            status,
            "blocker":           extra_notes if status in ("Blocked", "On Hold", "Waiting", "At Risk") else None,
            "deployment_type":   str(deploy_type).strip(),
            "timeline_label":    timeline_label,
            "target_date":       target_date,
            "target_confidence": target_confidence,
            "hardware":          str(hardware).strip(),
            "estimated_cost":    cost,
            "team_owner":        owner,
            "devops_id":         devops,
            "notes":             notes.strip() if notes else None,
            "last_updated":      "2026-04-15T00:00:00",  # seeded from Excel snapshot
            "last_updated_by":   "seed_import",
        }

        create_project(project_data)

        # Create initial history entry
        add_history(
            project_id=pid,
            changes={"_seed": {"old": None, "new": "Imported from Excel"}},
            source_text=f"Seeded from SPARROW_BACKLOG_OF_PRIORITIES.xlsx row {row_num}",
            source_type="system",
            updated_by="seed_import",
            llm_summary=f"Initial import: {location} ({partner_org}) — {status}",
        )

        seeded += 1
        print(f"  {pid:20s}  {status:12s}  {location} ({partner_org})")

    print(f"\nSeeded {seeded} projects into {os.path.basename(os.path.abspath(xlsx_path))}.")
    print(f"Database: {os.path.abspath(os.path.join(os.path.dirname(__file__), 'sparrow_tracker.db'))}")


if __name__ == "__main__":
    seed()
